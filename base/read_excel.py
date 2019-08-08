from openpyxl import load_workbook
from base.read_config import get_config
def read_data(sheet_name,case_id):
    # 打开excel
    workbook1=load_workbook(get_config('DATABASE', 'data_address'))
    # otherr（test_data）
    sheet1=workbook1[sheet_name]
    print(sheet1)
    test_case=[]  #用来存储每一行数据，也就是一条测试用例
    test_case.append(sheet1.cell(case_id+1,1).value)
    test_case.append(sheet1.cell(case_id+1,2).value)
    test_case.append(sheet1.cell(case_id+1,3).value)
    test_case.append(sheet1.cell(case_id+1,4).value)
    test_case.append(sheet1.cell(case_id+1,5).value)
    test_case.append(sheet1.cell(case_id+1,6).value)
    test_case.append(sheet1.cell(case_id+1,7).value)
    test_case.append(sheet1.cell(case_id+1,8).value)
    test_case.append(sheet1.cell(case_id+1,9).value)
    return test_case    #将读取到的用例返回