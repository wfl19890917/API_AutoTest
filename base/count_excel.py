from openpyxl import load_workbook
from base.read_config import get_config
def count_case(sheet_name):
    workbook1=load_workbook(get_config('DATABASE', 'data_address'))
    sheet=workbook1[sheet_name]
    max_row=sheet.max_row  #统计测试用例的行数
    return max_row