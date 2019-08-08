from openpyxl import load_workbook
from openpyxl.styles import Font,colors,PatternFill
from base.read_config import get_config
import shutil
import os

def write_data(sheet_name,row,col,value,red=False):
    workbook1=load_workbook(get_config('DATABASE', 'data_address'))
    sheet=workbook1[sheet_name]
    sheet.cell(row,col).value=value
    if red==False:
        pass
    else:
        sheet.cell(row,col).font=Font(color=colors.RED)
        #sheet.cell(row,col).fill=PatternFill("solid", fgColor="FF0000")
    workbook1.save(get_config('DATABASE', 'data_address'))
    #newworkbook.save(get_config('DATABASE', 'data_address'))
